from nsepy import get_history as gh
from datetime import date
from openpyxl import Workbook
import pandas as pd
import csv
import datetime as dt
import dateutil.relativedelta as dr
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\vijay\Downloads\List of Stocks case study.xlsx")  
ws = wb["Sheet2"] 
column = ws['B']  
tickers = [column[x].value for x in range(1,len(column))]
end=dt.date.today()
start=end-dr.relativedelta(days=90)
GFG = pd.ExcelWriter('finaloutput.xlsx')
for ticker in tickers:
    t1=ticker
    t1=t1.replace(" ", "")
    try:
        data=gh(t1,start,end)
        df = data['Close']
        df.to_csv("cache.csv")
        df_new = pd.read_csv('cache.csv')
        df_new.to_excel(GFG,sheet_name=ticker, index = False)
    except:
        print("error in "+ticker)
GFG.save()
